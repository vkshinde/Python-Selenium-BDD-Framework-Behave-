import logging
import time

from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementClickInterceptedException, ElementNotVisibleException, TimeoutException, \
    NoSuchElementException, ElementNotInteractableException, InvalidElementStateException, \
    InvalidSelectorException as EX
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException

from openpyxl import *

import os

current_path = os.getcwd()


# print(current_path)


class BasePage:
    __value = 0
    def __init__(self, driver, file="Test_required_data/FlipkartTesting.xlsx" ):
        self.driver = driver
        self.__value = 0
        self.file = file

        self.__excel = ReadExcel(self.file)

    def print_cdr(self):
        print(os.getcwd())

    def click_through_JS(self, by_locator):
        element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator))
        self.driver.execute_script(f"searchAutoCompleteDropDownHomePageClick({element})")

    def ajax_click_click_element(self, by_locator):
        element = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(by_locator))
        self.driver.execute_script("arguments[0].click();", element)

    def return_list(self, locator):
        return self.driver.find_elements_by_xpath(locator)

    def click_element(self, by_locator):
        try:
            # element = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((by_locator)))
            element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator))
            self.driver.execute_script("arguments[0].click();", element)
        except TimeoutException:
            logging.info(str(by_locator) + "Element not found.")
            return None


    def action_Chains_Element(self, by_locator):
        element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator))
        actions = ActionChains(self.driver)
        actions.click(element)
        actions.perform()
    def action_Chains_Click(self):
        actions = ActionChains(self.driver)
        actions.move_by_offset(100, 100)
        actions.click()
        actions.perform()

    def execute_js(self, test):

        self.driver.execute_script(test)

    def return_execute_js(self, test):
        return self.driver.execute_script(test)

    def input_element(self, by_locator, text):
        try:

            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator)).send_keys(text)
        except EX as e:
            print("Exception! Can't click on the element")

    def switch_to_content(self):

        self.driver.switch_to.default_content()

    def takeScreenShotOfThePage(self):

        self.__value += 1
        self.driver.save_screenshot("ScreenShots//screenie" + str(self.__value) + '.png')

    def get_element_text(self, by_locator):
        element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator))
        return element.text

    def send_Enter_key(self, by_locator):
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator)).send_keys(Keys.ENTER)

    def switch_to_pop_up_and_accept(self):
        alert = self.driver.switch_to.alert
        alert.accept()

    def get_title(self):
        return self.driver.title

    def get_current_url(self):
        return self.driver.current_url

    def get_element_attribute(self, by_locator, attribute_name):
        element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator))
        return element.get_attribute(attribute_name)

    def presence_of_element_located_dev(self, by_locator):
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((by_locator)))

    # you can use this function to check the specific element is present on the webpage or not
    def verify_element_displayed(self, by_locator):
        try:

            element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator))
            return element.is_displayed()
        except:
            return False

    # If there is an Iframe in your code and you want to access the elements of it
    # then you can use this function to
    def switch_to_iframe(self, locator):
        iframe = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(locator))
        self.driver.switch_to.frame(iframe)

    # This function is used for refreshing the page
    def refresh_the_driver(self):
        self.driver.refresh()

    # This function will be used to load a new url
    def load_url(self, url):
        t = time.time()
        self.driver.set_page_load_timeout(10)
        try:
            self.driver.get(url)
        except TimeoutException:
            self.driver.execute_script("window.stop();")

    def read_data(self, sheet_name):

        return self.__excel.read_excel(sheet_name)



class ReadExcel:
    def __init__(self, excel_file_path):
        self.dict = dict()
        self.list = list()
        self.key = ""
        self.file = excel_file_path
        # print(current_path)
        # os.chdir("../")
        try:
            self.wb = load_workbook(self.file)
        except FileNotFoundError:
            logging.log(msg="File not found")
        self.sheet = ""

    def __add(self, key, value):
        self.dict[key] = value

    def read_excel(self, sheet_name):


            try:
                self.sheet = self.wb[sheet_name]
            except KeyError:
                print(sheet_name + " -- sheet is not present in the sheet")
            for column in range(1, self.sheet.max_column + 1):
                for row in range(1, self.sheet.max_row + 1):
                    if row == 1:
                        self.list.clear()
                        self.key = self.sheet.cell(row, column).value
                        self.__add(self.key, list(self.list))
                    else:
                        self.list.append(self.sheet.cell(row, column).value)
                        self.__add(self.key, list(self.list))

            os.chdir(current_path)
            return self.dict

    def print_dict(self):
        print(self.dict)

